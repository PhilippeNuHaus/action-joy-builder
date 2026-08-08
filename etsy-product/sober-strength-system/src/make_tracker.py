#!/usr/bin/env python3
"""Builds Sober-Strength-Tracker.xlsx — the editable companion to the PDF program.

Deliberately formula-driven rather than a static grid: the sober-day count, the
Green/Amber/Red status, the estimated 1RM and every dashboard figure calculate
themselves, so the buyer types the minimum and the sheet does the rest.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

INK = "12161C"
ACCENT = "C2410C"
PANEL = "F4F4F5"
LINE = "D8DAE0"
GREEN = "15803D"
AMBER = "B45309"
RED = "B91C1C"

HEAD = Font(name="Arial", size=9, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=16, bold=True, color=INK)
SUB = Font(name="Arial", size=10, color="3D4552")
BODY = Font(name="Arial", size=10, color="3D4552")
BOLD = Font(name="Arial", size=10, bold=True, color=INK)
ACC = Font(name="Arial", size=9, bold=True, color=ACCENT)
FINE = Font(name="Arial", size=8, color="6B7280")

FILL_HEAD = PatternFill("solid", fgColor=INK)
FILL_PANEL = PatternFill("solid", fgColor=PANEL)
FILL_INPUT = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color=LINE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

ROWS = 100          # check-in rows: 90 days plus buffer
LOG_ROWS = 420      # training log rows: 44 sessions x ~6 exercises plus buffer


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def header_row(ws, row, labels, start=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=label)
        c.font = HEAD
        c.fill = FILL_HEAD
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX
    ws.row_dimensions[row].height = 20


def title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws["A2"] = subtitle
    ws["A2"].font = SUB
    ws.row_dimensions[1].height = 22


def build():
    wb = Workbook()

    # ------------------------------------------------------------------ lists
    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    for i, v in enumerate(["Good", "Broken", "Barely"], start=1):
        lists.cell(row=i, column=1, value=v)
    for i, v in enumerate(["Quiet", "Noticeable", "Loud"], start=1):
        lists.cell(row=i, column=2, value=v)
    for i, v in enumerate(["Yes", "No", "Red Session"], start=1):
        lists.cell(row=i, column=3, value=v)
    phases = ["01 Anchor", "02 Build", "03 Claim"]
    for i, v in enumerate(phases, start=1):
        lists.cell(row=i, column=4, value=v)
    sessions = ["A", "B", "C", "Upper A", "Lower A", "Upper B", "Lower B", "Test"]
    for i, v in enumerate(sessions, start=1):
        lists.cell(row=i, column=5, value=v)

    # ----------------------------------------------------------------- readme
    ws = wb.active
    ws.title = "Read Me"
    widths(ws, {"A": 3, "B": 92})
    ws["B1"] = "THE SOBER STRENGTH SYSTEM"
    ws["B1"].font = TITLE
    ws["B2"] = "Editable tracker — companion to the 12-week program PDF"
    ws["B2"].font = SUB
    ws.row_dimensions[1].height = 22

    lines = [
        ("", ""),
        ("HOW TO USE THIS FILE", "h"),
        ("1.  Open the 'Daily Check-In' tab and type your start date in cell B4. Everything else dates itself from there.", "b"),
        ("2.  Each morning, fill in the Sleep and Craving columns from the dropdowns. The Status column turns itself", "b"),
        ("     Green, Amber or Red and tells you which version of the session to run.", "b"),
        ("3.  After training, add your sets to the 'Training Log' tab. Estimated 1RM calculates automatically.", "b"),
        ("4.  The 'Dashboard' tab totals everything. You never type anything into it.", "b"),
        ("", ""),
        ("WHICH CELLS YOU TYPE IN", "h"),
        ("White cells with a border are for you. Grey cells are calculated — if you type over one you will break", "b"),
        ("its formula. If that happens, close the file without saving and reopen it.", "b"),
        ("", ""),
        ("OPENS IN", "h"),
        ("Microsoft Excel, Apple Numbers, Google Sheets (File > Import > Upload), and LibreOffice Calc, which is free.", "b"),
        ("Formulas are standard and work in all four. On Google Sheets, import as a new spreadsheet rather than", "b"),
        ("opening in preview mode, or the dropdowns will not appear.", "b"),
        ("", ""),
        ("IMPORTANT — PLEASE READ", "w"),
        ("Stopping alcohol suddenly can be medically dangerous if you drink heavily or daily, and in some cases", "b"),
        ("fatal. This is a training tracker, not a detox or treatment plan. Speak to a doctor before you stop", "b"),
        ("drinking. In the US, SAMHSA's National Helpline is 1-800-662-4357 — free, confidential, 24/7.", "b"),
        ("", ""),
        ("This file is for general educational purposes only and is not medical advice. It does not replace care from a", "f"),
        ("doctor, therapist or addiction professional. Consult a qualified healthcare professional before beginning any", "f"),
        ("exercise program or changing alcohol consumption. Personal use only — not for resale or redistribution.", "f"),
    ]
    r = 4
    for text, kind in lines:
        c = ws.cell(row=r, column=2, value=text)
        if kind == "h":
            c.font = ACC
        elif kind == "w":
            c.font = Font(name="Arial", size=9, bold=True, color=RED)
        elif kind == "f":
            c.font = FINE
        else:
            c.font = BODY
        r += 1

    # ----------------------------------------------------------- daily check-in
    ci = wb.create_sheet("Daily Check-In")
    widths(ci, {"A": 12, "B": 11, "C": 13, "D": 14, "E": 12, "F": 15, "G": 46})
    ci["A1"] = "DAILY CHECK-IN"
    ci["A1"].font = TITLE
    ci["A2"] = "Sixty seconds every morning. The Status column picks your session — see page 7 of the program."
    ci["A2"].font = SUB

    ci["A4"] = "Start date"
    ci["A4"].font = BOLD
    ci["B4"] = "=TODAY()"
    ci["B4"].font = BOLD
    ci["B4"].number_format = "yyyy-mm-dd"
    ci["B4"].fill = FILL_INPUT
    ci["B4"].border = BOX
    ci["C4"] = "← type your day 1 here"
    ci["C4"].font = FINE

    header_row(ci, 6, ["Date", "Sober day", "Sleep", "Craving", "Status", "Session done", "Notes"])
    ci.freeze_panes = "A7"

    dv_sleep = DataValidation(type="list", formula1="=Lists!$A$1:$A$3", allow_blank=True)
    dv_crave = DataValidation(type="list", formula1="=Lists!$B$1:$B$3", allow_blank=True)
    dv_done = DataValidation(type="list", formula1="=Lists!$C$1:$C$3", allow_blank=True)
    for dv in (dv_sleep, dv_crave, dv_done):
        ci.add_data_validation(dv)

    for i in range(ROWS):
        r = 7 + i
        ci.cell(row=r, column=1, value=f"=$B$4+{i}").number_format = "yyyy-mm-dd"
        ci.cell(row=r, column=2, value=f"=A{r}-$B$4+1")
        # Status: Red dominates, then Amber, then Green.
        ci.cell(
            row=r,
            column=5,
            value=(
                f'=IF(OR(C{r}="",D{r}=""),"",'
                f'IF(OR(C{r}="Barely",D{r}="Loud"),"RED",'
                f'IF(OR(C{r}="Broken",D{r}="Noticeable"),"AMBER","GREEN")))'
            ),
        )
        for col in range(1, 8):
            cell = ci.cell(row=r, column=col)
            cell.border = BOX
            cell.font = BODY
            if col in (1, 2, 5):
                cell.fill = FILL_PANEL
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.fill = FILL_INPUT
        ci.cell(row=r, column=5).font = Font(name="Arial", size=9, bold=True)
        dv_sleep.add(ci.cell(row=r, column=3))
        dv_crave.add(ci.cell(row=r, column=4))
        dv_done.add(ci.cell(row=r, column=6))

    rng = f"E7:E{6 + ROWS}"
    ci.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"GREEN"'],
        fill=PatternFill("solid", fgColor="DCFCE7"), font=Font(name="Arial", size=9, bold=True, color=GREEN)))
    ci.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"AMBER"'],
        fill=PatternFill("solid", fgColor="FEF3C7"), font=Font(name="Arial", size=9, bold=True, color=AMBER)))
    ci.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"RED"'],
        fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(name="Arial", size=9, bold=True, color=RED)))

    # -------------------------------------------------------------- training log
    tl = wb.create_sheet("Training Log")
    widths(tl, {"A": 12, "B": 11, "C": 12, "D": 11, "E": 30, "F": 11, "G": 9, "H": 8, "I": 13, "J": 30})
    tl["A1"] = "TRAINING LOG"
    tl["A1"].font = TITLE
    tl["A2"] = "One row per set, or one row per exercise using your top set. Estimated 1RM calculates itself."
    tl["A2"].font = SUB

    header_row(tl, 4, ["Date", "Sober day", "Phase", "Session", "Exercise", "Weight", "Reps",
                       "RPE", "Est. 1RM", "Notes"])
    tl.freeze_panes = "A5"

    dv_phase = DataValidation(type="list", formula1="=Lists!$D$1:$D$3", allow_blank=True)
    dv_sess = DataValidation(type="list", formula1="=Lists!$E$1:$E$8", allow_blank=True)
    for dv in (dv_phase, dv_sess):
        tl.add_data_validation(dv)

    for i in range(LOG_ROWS):
        r = 5 + i
        tl.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        # Sober day derives from the check-in start date, so it is never typed twice.
        tl.cell(row=r, column=2, value=f"=IF(A{r}=\"\",\"\",A{r}-'Daily Check-In'!$B$4+1)")
        # Epley estimate; a single rep is already the maximum so it passes through.
        tl.cell(
            row=r,
            column=9,
            value=f'=IF(OR(F{r}="",G{r}=""),"",ROUND(F{r}*(1+G{r}/30),1))',
        )
        for col in range(1, 11):
            cell = tl.cell(row=r, column=col)
            cell.border = BOX
            cell.font = BODY
            if col in (2, 9):
                cell.fill = FILL_PANEL
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.fill = FILL_INPUT
            if col in (6, 7, 8):
                cell.alignment = Alignment(horizontal="center")
        dv_phase.add(tl.cell(row=r, column=3))
        dv_sess.add(tl.cell(row=r, column=4))

    # ----------------------------------------------------------------- dashboard
    db = wb.create_sheet("Dashboard")
    widths(db, {"A": 3, "B": 34, "C": 16, "D": 4, "E": 34, "F": 16})
    db["B1"] = "DASHBOARD"
    db["B1"].font = TITLE
    db["B2"] = "Everything here calculates itself. Nothing on this tab needs typing."
    db["B2"].font = SUB

    ci_end = 6 + ROWS
    tl_end = 4 + LOG_ROWS

    def block(anchor_col, row0, heading, rows):
        c = db.cell(row=row0, column=anchor_col, value=heading)
        c.font = ACC
        rr = row0 + 1
        for label, formula, fmt in rows:
            lc = db.cell(row=rr, column=anchor_col, value=label)
            lc.font = BODY
            lc.border = BOX
            lc.fill = FILL_PANEL
            vc = db.cell(row=rr, column=anchor_col + 1, value=formula)
            vc.font = BOLD
            vc.border = BOX
            vc.fill = FILL_PANEL
            vc.alignment = Alignment(horizontal="center")
            if fmt:
                vc.number_format = fmt
            rr += 1
        return rr

    # Counting the three status words is deliberate: the Status column holds formulas
    # that return "", which COUNTIF(range,"<>") would wrongly count as filled.
    filled = (f"COUNTIF('Daily Check-In'!E7:E{ci_end},\"GREEN\")"
              f"+COUNTIF('Daily Check-In'!E7:E{ci_end},\"AMBER\")"
              f"+COUNTIF('Daily Check-In'!E7:E{ci_end},\"RED\")")
    sobriety = [
        ("Current sober day", f"=IFERROR(MAX(0,TODAY()-'Daily Check-In'!$B$4+1),0)", "0"),
        ("Days checked in", f"={filled}", "0"),
        ("Green days", f"=COUNTIF('Daily Check-In'!E7:E{ci_end},\"GREEN\")", "0"),
        ("Amber days", f"=COUNTIF('Daily Check-In'!E7:E{ci_end},\"AMBER\")", "0"),
        ("Red days", f"=COUNTIF('Daily Check-In'!E7:E{ci_end},\"RED\")", "0"),
        ("Green day rate",
         f"=IFERROR(COUNTIF('Daily Check-In'!E7:E{ci_end},\"GREEN\")/({filled}),0)",
         "0%"),
    ]
    training = [
        ("Sessions completed", f"=COUNTIF('Daily Check-In'!F7:F{ci_end},\"Yes\")", "0"),
        ("Red Sessions completed", f"=COUNTIF('Daily Check-In'!F7:F{ci_end},\"Red Session\")", "0"),
        ("Total sessions (both count)",
         f"=COUNTIF('Daily Check-In'!F7:F{ci_end},\"Yes\")+COUNTIF('Daily Check-In'!F7:F{ci_end},\"Red Session\")", "0"),
        ("Sets logged", f"=COUNT('Training Log'!F5:F{tl_end})", "0"),
        ("Total volume (weight x reps)",
         f"=IFERROR(SUMPRODUCT('Training Log'!F5:F{tl_end},'Training Log'!G5:G{tl_end}),0)", "#,##0"),
        ("Program adherence (of 44)",
         f"=IFERROR((COUNTIF('Daily Check-In'!F7:F{ci_end},\"Yes\")"
         f"+COUNTIF('Daily Check-In'!F7:F{ci_end},\"Red Session\"))/44,0)", "0%"),
    ]

    next_row = block(2, 4, "SOBRIETY", sobriety)
    block(5, 4, "TRAINING", training)

    # Best estimated 1RM per main lift, matched on any exercise name containing the keyword.
    lifts = [("Squat", "squat"), ("Bench", "bench"), ("Deadlift", "deadlift"),
             ("Overhead Press", "overhead"), ("Row", "row"), ("Pull-Up / Pulldown", "pull")]
    hdr = db.cell(row=next_row + 1, column=2, value="BEST ESTIMATED 1RM")
    hdr.font = ACC
    note = db.cell(row=next_row + 1, column=5, value="Matches any exercise name containing the keyword.")
    note.font = FINE
    rr = next_row + 2
    for label, key in lifts:
        lc = db.cell(row=rr, column=2, value=label)
        lc.font = BODY
        lc.border = BOX
        lc.fill = FILL_PANEL
        vc = db.cell(
            row=rr, column=3,
            value=f'=IFERROR(MAX(IF(ISNUMBER(SEARCH("{key}",\'Training Log\'!$E$5:$E${tl_end})),'
                  f'\'Training Log\'!$I$5:$I${tl_end})),0)',
        )
        vc.font = BOLD
        vc.border = BOX
        vc.fill = FILL_PANEL
        vc.alignment = Alignment(horizontal="center")
        vc.number_format = "0.0"
        rr += 1

    n = db.cell(row=rr + 1, column=2,
                value="These six are array formulas. In Excel 2019 or older, click the cell, press F2 then "
                      "Ctrl+Shift+Enter. Excel 365, Google Sheets, Numbers and LibreOffice need no extra step.")
    n.font = FINE

    d = db.cell(row=rr + 3, column=2,
                value="Not medical advice. Alcohol withdrawal can be dangerous — speak to a doctor. "
                      "SAMHSA National Helpline (US): 1-800-662-4357.")
    d.font = FINE

    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "build")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "Sober-Strength-Tracker.xlsx")
    wb.save(path)
    print(f"Wrote {path}")
    return path


if __name__ == "__main__":
    build()
